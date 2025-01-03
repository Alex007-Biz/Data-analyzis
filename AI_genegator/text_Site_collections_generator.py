import openpyxl
import openai
from config import api_key
from openai import OpenAI


# Установите ваш API-ключ OpenAI
openai.api_key = api_key

excel_path = "collections_try.xlsx"

client = OpenAI(
    api_key=api_key,  # This is the default and can be omitted
)

# Функция для получения описания от ChatGPT
def get_description(brand, collection):
    message = 'Напиши небольшое привлекательное описание для коллекции плитки ' + collection + ' фабрики ' + brand
    print(message)
    content = 'Вы креативный копирайтер. Найди в интернете технические характеристики и фото этого товара. Проанализируй фото и информацию о товаре, выдели ключевые характеристики и преимущества. Проанализируй стиль, цвет, форму, поверхность, размеры товара, его художественное исполнение. Твое описание товара должно быть яркими и привлекать внимание. Добавь уникальности, пиши как топовый маркетолог, внеси в текст интересные факты. Используй html-разметку. Описание должно быть на 100% уникальное!'
    try:
        chat_completion = client.chat.completions.create(
            model="chatgpt-4o-latest",
            max_tokens=600,
            temperature=0.7,
            messages=[
                {
                    "role": "system",
                    "content": content,
                },
                {
                    "role": "user",
                    "content": message,
                }
            ]
        )
        # Извлекаем описание из ответа
        return chat_completion.choices[0].message.content.strip()
    except Exception as e:
        raise RuntimeError(f"Ошибка при обработке запроса для {product}: {e}")

# Открываем Excel-файл
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active
i = 0

try:
    # Проход по строкам
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=6):
        i += 1
        # product = row[2].value  # Значение из столбца "товар"
        brand = row[2].value  # Значение из столбца "Бренд"
        if isinstance(brand, str) and brand:
            brand = brand.capitalize()
        else:
            brand = ""
        collection = row[3].value  # Значение из столбца "Коллекция"
        if isinstance(collection, str) and collection:
            collection = collection.capitalize()
        else:
            collection = ""

        if product:
            print(f"{i}. Обрабатываю: товар - {product}")
            try:
                description = get_description(product, brand, collection)
            except RuntimeError as e:
                print(e)
                break  # Прерываем цикл при ошибке
            row[4].value = description  # Записываем описание в столбец "Описание"

finally:
    # Сохраняем изменения в Excel-файле независимо от результата
    workbook.save(excel_path)
    print(f"Описание сохранено в {excel_path}.")
